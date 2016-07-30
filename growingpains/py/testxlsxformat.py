from openpyxl import load_workbook
wb = load_workbook(filename='dbfile.xlsx', read_only=True)


excludetabs = ['目录', 'MARK']
DEType = ['土建', '安装', '拆除', '房修']
materials = ['人工', '材料', '机械']

# 数据行值
def rowdata(row):
    vals = [];
    for cell in row:
        vals.append(cell.value)
    return vals


def printToFileOfRow(filename):
    # tabs = wb.get_sheet_names()
    tabs = [i for i in wb.get_sheet_names() if i not in excludetabs]
    # print to file
    f = open(filename,'w')
    for name in tabs:
        print(tabs.index(name) + 1, '/',  len(tabs) , ' ', name)
        print('#%s' % name, file=f)
        sheet = wb[name]
        for row in sheet.rows:
            print(rowdata(row), file=f)    
    f.close()


printToFileOfRow('basedata.txt')







ws = wb['MARK']
ws = wb['DEINDEX']



print(wb.get_sheet_names())
# ws = wb.active # ws is now an IterableWorksheet

def printtablevalues(table):
    for row in table.rows:
        for cell in row:
            print(cell.value)



for row in ws.iter_rows('A1:C2'):
    for cell in row:
        print (cell.value)


i = 0
for row in ws.rows:
    i = i + 1
    if i > 2 and i < 5:

        print(rowdata(row))



for row in ws.iter_rows('A1:C2'):
    print (rowdata(row))


def writesheet(f, sheet):
    for row in sheet.rows:
        writerowdata(f, row)


for row in ws.rows:
    print(rowdata(row), file=f)





excludes = ['目录', 'MARK']
DEType = ['土建', '安装', '拆除', '房修']
materials = ['人工', '材料', '机械']



list3=[i for i in list1 if i not in list2]


# 
ws = wb['DEINDEX']
for i in range(ws.min_row + 2, ws.max_row):
    if i >= 100 and i % 100 == 0:
        print(i, '/' , ws.max_row)
    dtname = ws.cell(row = i, column = 2).value
    if not dtname in DEType:
        DEType.append(dtname)
        print(dtname)


for i in range(ws.min_row + 2, ws.max_row):
    if i >= 100 and i % 500 == 0:
        print(i, '/' , ws.max_row)


excludetabs = ['目录', 'MARK']
DEType = ['土建', '安装', '拆除', '房修']
materials = ['人工', '材料', '机械']

# 数据行值
def rowdata(row):
    vals = [];
    for cell in row:
        vals.append(cell.value)
    return vals


def printToFileOfRow(filename):
    # tabs = wb.get_sheet_names()
    tabs = [i for i in wb.get_sheet_names() if i not in excludetabs]
    # print to file
    f = open(filename,'w')
    for name in tabs:
        print(tabs.index(name), '/',  len(tabs) , ' ', name)
        print('#%s' % name, file=f)
        sheet = wb[name]
        for row in sheet.rows:
            print(rowdata(row), file=f)    
    f.close()


printToFileOfRow('basedata.txt')




# 输出行数据列表
def rowdatalist(ws, r):
    vals = []
    for c in range(ws.min_column, ws.max_column):
        vals.append(ws.cell(row = r, column = c).value)
    return vals   

def printtofile(filename):
    # tabs = wb.get_sheet_names()
    tabs = [i for i in wb.get_sheet_names() if i not in excludetabs]
    f = open(filename,'w')
    for name in tabs:
        print(tabs.index(9), '/',  tabs.count() , ' ', name)
        print('#%s' % name, file=f)
        sheet = wb[name]
        # title row
        print(rowdatalist(sheet, 1), file=f)
        # data row
        for r in range(ws.min_row + 2, ws.max_row):
            # 显示进度行
            if r >= 100 and r % 100 == 0:
                print('  ' + r, '/' , ws.max_row) 
            print(rowdatalist(sheet, r), file=f)
    f.close() 

